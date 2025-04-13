import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from mpl_toolkits.mplot3d import Axes3D

# 定义算法名称列表
algorithms = [
    "Average Hash (aHash)",
    "Perceptual Hash (pHash)",
    "Difference Hash (dHash)",
    "Wavelet Hash (wHash)",
    "Color Hash (cHash)",
    "Marr-Hildreth Hash (mhHash)",
    "ResNet50 Hash",
    "ResNet50 Deep Features",
    "Multiscale ResNet50 Hash",
    "Multiscale ResNet50 Deep Features",
    "ViT Hash",
    "ViT Deep Features",
    "Multiscale ViT Hash",
    "Multiscale ViT Deep Features",
    "Saliency-Enhanced ResNet Hash",
    "Saliency-Enhanced ResNet Deep Features"
]

# 定义哈希大小列表
hash_sizes = [4, 8, 16, 32, 64, 128, 256, 1024]

# Copy2数据集
# mAP数据
copy2_map_data = [
    [0.4637, 0.6340, 0.6679, 0.6689, 0.6689, 0.6689, 0.6689, 0.6689],
    [0.3418, 0.5447, 0.5633, 0.5680, 0.5680, 0.5680, 0.5680, 0.5680],
    [0.4148, 0.5505, 0.5750, 0.5890, 0.5890, 0.5890, 0.5890, 0.5890],
    [0.3313, 0.5534, 0.5866, 0.5997, 0.5997, 0.5997, 0.5997, 0.5997],
    [0.0547, 0.4042, 0.4698, 0.3585, 0.3585, 0.3585, 0.3585, 0.3585],
    [0.3568, 0.5094, 0.5392, 0.5491, 0.5491, 0.5491, 0.5491, 0.5491],
    [0.3274, 0.7560, 0.8617, 0.8906, 0.8906, 0.8906, 0.8906, 0.8906],
    [0.8814, 0.8814, 0.8814, 0.8814, 0.8814, 0.8814, 0.8814, 0.8814],
    [0.3770, 0.7794, 0.8708, 0.8953, 0.8953, 0.8953, 0.8953, 0.8953],
    [0.8881, 0.8881, 0.8881, 0.8881, 0.8881, 0.8881, 0.8881, 0.8881],
    [0.3925, 0.7742, 0.8741, 0.8927, 0.8927, 0.8927, 0.8927, 0.8927],
    [0.9012, 0.9012, 0.9012, 0.9012, 0.9012, 0.9012, 0.9012, 0.9012],
    [0.4169, 0.8005, 0.8899, 0.9103, 0.9103, 0.9103, 0.9103, 0.9103],
    [0.9146, 0.9146, 0.9146, 0.9146, 0.9146, 0.9146, 0.9146, 0.9146],
    [0.0483, 0.2919, 0.5465, 0.7092, 0.7092, 0.7092, 0.7092, 0.7092],
    [0.7671, 0.7671, 0.7671, 0.7671, 0.7671, 0.7671, 0.7671, 0.7671]
]

# μAP数据
copy2_uap_data = [
    [0.1382, 0.4695, 0.5985, 0.6008, 0.6008, 0.6008, 0.6008, 0.6008],
    [0.3291, 0.5429, 0.5619, 0.5671, 0.5671, 0.5671, 0.5671, 0.5671],
    [0.4057, 0.5481, 0.5752, 0.5853, 0.5853, 0.5853, 0.5853, 0.5853],
    [0.2502, 0.5325, 0.5682, 0.5767, 0.5767, 0.5767, 0.5767, 0.5767],
    [0.0137, 0.3268, 0.3915, 0.1742, 0.1742, 0.1742, 0.1742, 0.1742],
    [0.3123, 0.5014, 0.5302, 0.5359, 0.5359, 0.5359, 0.5359, 0.5359],
    [0.2837, 0.7166, 0.8296, 0.8584, 0.8584, 0.8584, 0.8584, 0.8584],
    [0.8479, 0.8479, 0.8479, 0.8479, 0.8479, 0.8479, 0.8479, 0.8479],
    [0.3380, 0.7403, 0.8377, 0.8636, 0.8636, 0.8636, 0.8636, 0.8636],
    [0.8524, 0.8524, 0.8524, 0.8524, 0.8524, 0.8524, 0.8524, 0.8524],
    [0.3541, 0.7373, 0.8168, 0.8366, 0.8366, 0.8366, 0.8366, 0.8366],
    [0.8490, 0.8490, 0.8490, 0.8490, 0.8490, 0.8490, 0.8490, 0.8490],
    [0.3724, 0.7781, 0.8521, 0.8736, 0.8736, 0.8736, 0.8736, 0.8736],
    [0.8765, 0.8765, 0.8765, 0.8765, 0.8765, 0.8765, 0.8765, 0.8765],
    [0.0184, 0.2092, 0.4990, 0.6003, 0.6003, 0.6003, 0.6003, 0.6003],
    [0.6826, 0.6826, 0.6826, 0.6826, 0.6826, 0.6826, 0.6826, 0.6826]
]

# Copydays数据集
# mAP数据
copydays_map_data = [
    [0.4449, 0.6368, 0.6717, 0.6682, 0.6682, 0.6682, 0.6682, 0.6682],
    [0.2849, 0.5589, 0.5727, 0.5347, 0.5347, 0.5347, 0.5347, 0.5347],
    [0.3495, 0.5561, 0.5721, 0.5514, 0.5514, 0.5514, 0.5514, 0.5514],
    [0.2576, 0.5796, 0.6203, 0.6298, 0.6298, 0.6298, 0.6298, 0.6298],
    [0.0634, 0.4008, 0.4668, 0.3712, 0.3712, 0.3712, 0.3712, 0.3712],
    [0.2501, 0.4433, 0.4795, 0.4785, 0.4785, 0.4785, 0.4785, 0.4785],
    [0.1969, 0.7121, 0.8662, 0.9015, 0.9015, 0.9015, 0.9015, 0.9015],
    [0.8893, 0.8893, 0.8893, 0.8893, 0.8893, 0.8893, 0.8893, 0.8893],
    [0.2533, 0.7757, 0.8994, 0.9253, 0.9253, 0.9253, 0.9253, 0.9253],
    [0.9188, 0.9188, 0.9188, 0.9188, 0.9188, 0.9188, 0.9188, 0.9188],
    [0.2379, 0.7432, 0.8761, 0.9020, 0.9020, 0.9020, 0.9020, 0.9020],
    [0.9138, 0.9138, 0.9138, 0.9138, 0.9138, 0.9138, 0.9138, 0.9138],
    [0.2961, 0.8129, 0.9224, 0.9433, 0.9433, 0.9433, 0.9433, 0.9433],
    [0.9472, 0.9472, 0.9472, 0.9472, 0.9472, 0.9472, 0.9472, 0.9472],
    [0.0259, 0.1448, 0.4538, 0.6734, 0.6734, 0.6734, 0.6734, 0.6734],
    [0.7404, 0.7404, 0.7404, 0.7404, 0.7404, 0.7404, 0.7404, 0.7404]
]

# μAP数据
copydays_uap_data = [
    [0.1159, 0.3847, 0.5273, 0.5201, 0.5201, 0.5201, 0.5201, 0.5201],
    [0.2632, 0.5554, 0.5698, 0.5327, 0.5327, 0.5327, 0.5327, 0.5327],
    [0.3335, 0.5526, 0.5697, 0.5328, 0.5328, 0.5328, 0.5328, 0.5328],
    [0.1766, 0.5372, 0.5772, 0.5691, 0.5691, 0.5691, 0.5691, 0.5691],
    [0.0138, 0.2819, 0.3423, 0.1588, 0.1588, 0.1588, 0.1588, 0.1588],
    [0.2065, 0.4313, 0.4655, 0.4468, 0.4468, 0.4468, 0.4468, 0.4468],
    [0.1454, 0.6331, 0.8287, 0.8676, 0.8676, 0.8676, 0.8676, 0.8676],
    [0.8477, 0.8477, 0.8477, 0.8477, 0.8477, 0.8477, 0.8477, 0.8477],
    [0.2020, 0.7138, 0.8637, 0.8962, 0.8962, 0.8962, 0.8962, 0.8962],
    [0.8787, 0.8787, 0.8787, 0.8787, 0.8787, 0.8787, 0.8787, 0.8787],
    [0.1923, 0.6748, 0.7986, 0.8292, 0.8292, 0.8292, 0.8292, 0.8292],
    [0.8472, 0.8472, 0.8472, 0.8472, 0.8472, 0.8472, 0.8472, 0.8472],
    [0.2388, 0.7698, 0.8753, 0.9041, 0.9041, 0.9041, 0.9041, 0.9041],
    [0.9077, 0.9077, 0.9077, 0.9077, 0.9077, 0.9077, 0.9077, 0.9077],
    [0.0134, 0.0826, 0.3835, 0.5365, 0.5365, 0.5365, 0.5365, 0.5365],
    [0.6506, 0.6506, 0.6506, 0.6506, 0.6506, 0.6506, 0.6506, 0.6506]
]

# 定义数据集列表
datasets = ["Copy2", "Copydays"]
map_data_list = [copy2_map_data, copydays_map_data]
uap_data_list = [copy2_uap_data, copydays_uap_data]

# 为每个数据集生成Excel文件和图表
for i, dataset in enumerate(datasets):
    # 获取当前数据集的数据
    map_data = map_data_list[i]
    uap_data = uap_data_list[i]
    
    # 创建DataFrame
    map_df = pd.DataFrame(map_data, index=algorithms, columns=hash_sizes)
    uap_df = pd.DataFrame(uap_data, index=algorithms, columns=hash_sizes)
    
    # 创建Excel文件
    with pd.ExcelWriter(f'g:\\trae\\myhash\\{dataset.lower()}_results.xlsx', engine='openpyxl') as writer:
        # 写入mAP数据
        map_df.to_excel(writer, sheet_name='mAP')
        # 写入μAP数据
        uap_df.to_excel(writer, sheet_name='μAP')
        
        # 获取工作簿和工作表
        workbook = writer.book
        
        # 为每个表格添加格式
        for sheet_name in ['mAP', 'μAP']:
            worksheet = writer.sheets[sheet_name]
            
            # 设置列宽
            for j, col in enumerate(worksheet.columns):
                column_letter = get_column_letter(j + 1)
                if j == 0:
                    worksheet.column_dimensions[column_letter].width = 30
                else:
                    worksheet.column_dimensions[column_letter].width = 12
            
            # 设置标题行格式
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            header_font = Font(bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # 设置算法名称列格式
            for j in range(2, len(algorithms) + 2):
                cell = worksheet.cell(row=j, column=1)
                cell.font = Font(bold=True)
            
            # 为数据单元格添加边框和对齐方式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(algorithms) + 1, min_col=1, max_col=len(hash_sizes) + 1):
                for cell in row:
                    cell.border = thin_border
                    if cell.column > 1 and cell.row > 1:  # 数据单元格
                        cell.alignment = Alignment(horizontal='center')
                        # 设置条件格式（颜色渐变）
                        if sheet_name == 'mAP':
                            value = map_data[cell.row - 2][cell.column - 2]
                        else:
                            value = uap_data[cell.row - 2][cell.column - 2]
                        
                        # 根据值设置背景色（值越高，颜色越深）
                        intensity = int(255 - min(value * 255, 255))
                        color = f"{intensity:02X}{intensity:02X}FF"  # 蓝色渐变
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # 创建折线图工作表
        chart_sheet = workbook.create_sheet(title='Charts')
        
        # 为mAP创建折线图
        map_chart = LineChart()
        map_chart.title = f"{dataset} - mAP vs Hash Size"
        map_chart.x_axis.title = "Hash Size"
        map_chart.y_axis.title = "mAP"
        
        # 添加数据
        map_data_ref = Reference(workbook['mAP'], min_row=1, max_row=len(algorithms) + 1, min_col=2, max_col=len(hash_sizes) + 1)
        map_cats = Reference(workbook['mAP'], min_row=2, max_row=len(algorithms) + 1, min_col=1, max_col=1)
        map_chart.add_data(map_data_ref, titles_from_data=True)
        map_chart.set_categories(map_cats)
        
        # 设置图表大小
        map_chart.width = 30
        map_chart.height = 15
        
        # 添加图表到工作表
        chart_sheet.add_chart(map_chart, "A1")
        
        # 为μAP创建折线图
        uap_chart = LineChart()
        uap_chart.title = f"{dataset} - μAP vs Hash Size"
        uap_chart.x_axis.title = "Hash Size"
        uap_chart.y_axis.title = "μAP"
        
        # 添加数据
        uap_data_ref = Reference(workbook['μAP'], min_row=1, max_row=len(algorithms) + 1, min_col=2, max_col=len(hash_sizes) + 1)
        uap_cats = Reference(workbook['μAP'], min_row=2, max_row=len(algorithms) + 1, min_col=1, max_col=1)
        uap_chart.add_data(uap_data_ref, titles_from_data=True)
        uap_chart.set_categories(uap_cats)
        
        # 设置图表大小
        uap_chart.width = 30
        uap_chart.height = 15
        
        # 添加图表到工作表
        chart_sheet.add_chart(uap_chart, "A20")
    
    print(f"Excel文件已生成：g:\\trae\\myhash\\{dataset.lower()}_results.xlsx")
    
    # 创建可视化图表
    plt.figure(figsize=(15, 10))
    
    # 绘制mAP图表
    plt.subplot(2, 1, 1)
    for j, algo in enumerate(algorithms):
        plt.plot(hash_sizes, map_data[j], marker='o', label=algo)
    plt.xscale('log', base=2)
    plt.xlabel('Hash Size')
    plt.ylabel('mAP')
    plt.title(f'{dataset} - mAP vs Hash Size')
    plt.grid(True, alpha=0.3)
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    
    # 绘制μAP图表
    plt.subplot(2, 1, 2)
    for j, algo in enumerate(algorithms):
        plt.plot(hash_sizes, uap_data[j], marker='o', label=algo)
    plt.xscale('log', base=2)
    plt.xlabel('Hash Size')
    plt.ylabel('μAP')
    plt.title(f'{dataset} - μAP vs Hash Size')
    plt.grid(True, alpha=0.3)
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    
    plt.tight_layout()
    plt.savefig(f'g:\\trae\\myhash\\{dataset.lower()}_results_chart.png', dpi=300, bbox_inches='tight')
    print(f"图表已保存：g:\\trae\\myhash\\{dataset.lower()}_results_chart.png")
    
    # 创建三维柱状图
    # 准备数据网格
    X, Y = np.meshgrid(np.arange(len(hash_sizes)), np.arange(len(algorithms)))
    hash_size_labels = [str(size) for size in hash_sizes]
    algorithm_labels = [algo[:15] + '...' if len(algo) > 15 else algo for algo in algorithms]
    
    # 创建mAP的三维柱状图
    fig_map = plt.figure(figsize=(16, 10))
    ax_map = fig_map.add_subplot(111, projection='3d')
    dz = np.array(map_data).flatten()
    z_zeros = np.zeros_like(dz)
    dx = np.ones_like(dz) * 0.5
    dy = np.ones_like(dz) * 0.5
    
    # 设置颜色映射
    colors = plt.cm.viridis(dz / np.max(dz))
    
    # 绘制柱状图
    ax_map.bar3d(X.flatten(), Y.flatten(), z_zeros, dx, dy, dz, color=colors, shade=True, alpha=0.8)
    
    # 设置坐标轴标签和标题
    ax_map.set_xlabel('Hash Size')
    ax_map.set_ylabel('Algorithm')
    ax_map.set_zlabel('mAP')
    ax_map.set_title(f'{dataset} - 3D Bar Chart of mAP Values')
    
    # 设置坐标轴刻度标签
    ax_map.set_xticks(np.arange(len(hash_sizes)))
    ax_map.set_yticks(np.arange(len(algorithms)))
    ax_map.set_xticklabels(hash_size_labels)
    ax_map.set_yticklabels(algorithm_labels)
    
    # 调整视角
    ax_map.view_init(elev=30, azim=45)
    
    # 保存图表
    plt.tight_layout()
    plt.savefig(f'g:\\trae\\myhash\\{dataset.lower()}_results_3d_map.png', dpi=300, bbox_inches='tight')
    print(f"mAP三维柱状图已保存：g:\\trae\\myhash\\{dataset.lower()}_results_3d_map.png")
    
    # 创建μAP的三维柱状图
    fig_uap = plt.figure(figsize=(16, 10))
    ax_uap = fig_uap.add_subplot(111, projection='3d')
    dz = np.array(uap_data).flatten()
    z_zeros = np.zeros_like(dz)
    dx = np.ones_like(dz) * 0.5
    dy = np.ones_like(dz) * 0.5
    
    # 设置颜色映射
    colors = plt.cm.viridis(dz / np.max(dz))
    
    # 绘制柱状图
    ax_uap.bar3d(X.flatten(), Y.flatten(), z_zeros, dx, dy, dz, color=colors, shade=True, alpha=0.8)
    
    # 设置坐标轴标签和标题
    ax_uap.set_xlabel('Hash Size')
    ax_uap.set_ylabel('Algorithm')
    ax_uap.set_zlabel('μAP')
    ax_uap.set_title(f'{dataset} - 3D Bar Chart of μAP Values')
    
    # 设置坐标轴刻度标签
    ax_uap.set_xticks(np.arange(len(hash_sizes)))
    ax_uap.set_yticks(np.arange(len(algorithms)))
    ax_uap.set_xticklabels(hash_size_labels)
    ax_uap.set_yticklabels(algorithm_labels)
    
    # 调整视角
    ax_uap.view_init(elev=30, azim=45)
    
    # 保存图表
    plt.tight_layout()
    plt.savefig(f'g:\\trae\\myhash\\{dataset.lower()}_results_3d_uap.png', dpi=300, bbox_inches='tight')
    print(f"μAP三维柱状图已保存：g:\\trae\\myhash\\{dataset.lower()}_results_3d_uap.png")

print("所有数据集的Excel表格和图表生成完成！")
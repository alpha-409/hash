import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# 定义算法名称列表
algorithms = [
    "Average Hash (aHash)",
    "Perceptual Hash (pHash)",
    "Difference Hash (dHash)",
    "Wavelet Hash (wHash)",
    "Color Hash (cHash)",
    "Marr-Hildreth Hash (mhHash)",
    "ResNet50 Hash",
    "ResNet50 Deep Features",
    "Multiscale ResNet50 Hash",
    "Multiscale ResNet50 Deep Features",
    "ViT Hash",
    "ViT Deep Features",
    "Multiscale ViT Hash",
    "Multiscale ViT Deep Features",
    "Saliency-Enhanced ResNet Hash",
    "Saliency-Enhanced ResNet Deep Features"
]

# 定义哈希大小列表
hash_sizes = [4, 8, 16, 32, 64, 128, 256, 1024]

# mAP数据
map_data = [
    [0.6430, 0.7150, 0.7263, 0.7288, 0.7283, 0.7279, 0.7279, 0.7278],
    [0.3866, 0.5686, 0.5931, 0.5972, 0.5893, 0.5966, 0.5932, 0.5994],
    [0.4173, 0.5589, 0.5923, 0.6194, 0.6397, 0.6311, 0.6121, 0.5763],
    [0.3442, 0.5661, 0.6279, 0.6440, 0.6490, 0.6508, 0.6435, 0.6468],
    [0.1623, 0.4109, 0.5679, 0.5408, 0.5184, 0.5627, 0.5627, 0.5627],
    [0.4229, 0.5554, 0.5819, 0.5914, 0.5946, 0.5951, 0.5950, 0.3843],
    [0.4411, 0.7573, 0.8316, 0.8475, 0.8545, 0.8545, 0.8545, 0.8545],
    [0.8356, 0.8356, 0.8356, 0.8356, 0.8356, 0.8356, 0.8356, 0.8356],
    [0.4787, 0.7962, 0.8508, 0.8605, 0.8658, 0.8658, 0.8658, 0.8658],
    [0.8549, 0.8549, 0.8549, 0.8549, 0.8549, 0.8549, 0.8549, 0.8549],
    [0.4358, 0.7306, 0.8239, 0.8403, 0.8403, 0.8403, 0.8403, 0.8403],
    [0.8487, 0.8487, 0.8487, 0.8487, 0.8487, 0.8487, 0.8487, 0.8487],
    [0.3939, 0.7467, 0.8420, 0.8591, 0.8591, 0.8591, 0.8591, 0.8591],
    [0.8621, 0.8621, 0.8621, 0.8621, 0.8621, 0.8621, 0.8621, 0.8621],
    [0.0976, 0.3372, 0.6028, 0.7601, 0.7974, 0.8168, 0.8199, 0.8221],
    [0.7909, 0.7909, 0.7909, 0.7909, 0.7909, 0.7909, 0.7909, 0.7909]
]

# μAP数据
uap_data = [
    [0.6183, 0.6952, 0.7116, 0.7192, 0.7169, 0.7155, 0.7110, 0.7109],
    [0.3719, 0.5636, 0.5899, 0.5949, 0.5872, 0.5949, 0.5918, 0.5990],
    [0.3920, 0.5559, 0.5884, 0.6091, 0.6040, 0.5500, 0.4791, 0.4443],
    [0.3038, 0.5485, 0.6120, 0.6306, 0.6381, 0.6373, 0.6210, 0.6263],
    [0.0956, 0.2640, 0.4104, 0.3504, 0.3512, 0.4056, 0.4142, 0.4142],
    [0.3820, 0.5360, 0.5765, 0.5873, 0.5903, 0.5869, 0.5811, 0.1885],
    [0.3947, 0.7101, 0.8011, 0.8147, 0.8239, 0.8239, 0.8239, 0.8239],
    [0.7947, 0.7947, 0.7947, 0.7947, 0.7947, 0.7947, 0.7947, 0.7947],
    [0.4339, 0.7558, 0.8203, 0.8270, 0.8350, 0.8350, 0.8350, 0.8350],
    [0.8197, 0.8197, 0.8197, 0.8197, 0.8197, 0.8197, 0.8197, 0.8197],
    [0.3847, 0.6472, 0.7383, 0.7591, 0.7591, 0.7591, 0.7591, 0.7591],
    [0.7797, 0.7797, 0.7797, 0.7797, 0.7797, 0.7797, 0.7797, 0.7797],
    [0.2065, 0.6616, 0.7792, 0.8036, 0.8036, 0.8036, 0.8036, 0.8036],
    [0.8055, 0.8055, 0.8055, 0.8055, 0.8055, 0.8055, 0.8055, 0.8055],
    [0.0616, 0.2912, 0.5409, 0.6869, 0.7354, 0.7579, 0.7654, 0.7681],
    [0.7397, 0.7397, 0.7397, 0.7397, 0.7397, 0.7397, 0.7397, 0.7397]
]

# 创建DataFrame
map_df = pd.DataFrame(map_data, index=algorithms, columns=hash_sizes)
uap_df = pd.DataFrame(uap_data, index=algorithms, columns=hash_sizes)

# 创建Excel文件
with pd.ExcelWriter('g:/trae/myhash/scid_results.xlsx', engine='openpyxl') as writer:
    # 写入mAP数据
    map_df.to_excel(writer, sheet_name='mAP')
    # 写入μAP数据
    uap_df.to_excel(writer, sheet_name='μAP')
    
    # 获取工作簿和工作表
    workbook = writer.book
    
    # 为每个表格添加格式
    for sheet_name in ['mAP', 'μAP']:
        worksheet = writer.sheets[sheet_name]
        
        # 设置列宽
        for i, col in enumerate(worksheet.columns):
            column_letter = get_column_letter(i + 1)
            if i == 0:
                worksheet.column_dimensions[column_letter].width = 30
            else:
                worksheet.column_dimensions[column_letter].width = 12
        
        # 设置标题行格式
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 设置算法名称列格式
        for i in range(2, len(algorithms) + 2):
            cell = worksheet.cell(row=i, column=1)
            cell.font = Font(bold=True)
        
        # 为数据单元格添加边框和对齐方式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(algorithms) + 1, min_col=1, max_col=len(hash_sizes) + 1):
            for cell in row:
                cell.border = thin_border
                if cell.column > 1 and cell.row > 1:  # 数据单元格
                    cell.alignment = Alignment(horizontal='center')
                    # 设置条件格式（颜色渐变）
                    if sheet_name == 'mAP':
                        value = map_data[cell.row - 2][cell.column - 2]
                    else:
                        value = uap_data[cell.row - 2][cell.column - 2]
                    
                    # 根据值设置背景色（值越高，颜色越深）
                    intensity = int(255 - min(value * 255, 255))
                    color = f"{intensity:02X}{intensity:02X}FF"  # 蓝色渐变
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    # 创建折线图工作表
    chart_sheet = workbook.create_sheet(title='Charts')
    
    # 为mAP创建折线图
    map_chart = LineChart()
    map_chart.title = "mAP vs Hash Size"
    map_chart.x_axis.title = "Hash Size"
    map_chart.y_axis.title = "mAP"
    
    # 添加数据
    map_data_ref = Reference(workbook['mAP'], min_row=1, max_row=len(algorithms) + 1, min_col=2, max_col=len(hash_sizes) + 1)
    map_cats = Reference(workbook['mAP'], min_row=2, max_row=len(algorithms) + 1, min_col=1, max_col=1)
    map_chart.add_data(map_data_ref, titles_from_data=True)
    map_chart.set_categories(map_cats)
    
    # 设置图表大小
    map_chart.width = 30
    map_chart.height = 15
    
    # 添加图表到工作表
    chart_sheet.add_chart(map_chart, "A1")
    
    # 为μAP创建折线图
    uap_chart = LineChart()
    uap_chart.title = "μAP vs Hash Size"
    uap_chart.x_axis.title = "Hash Size"
    uap_chart.y_axis.title = "μAP"
    
    # 添加数据
    uap_data_ref = Reference(workbook['μAP'], min_row=1, max_row=len(algorithms) + 1, min_col=2, max_col=len(hash_sizes) + 1)
    uap_cats = Reference(workbook['μAP'], min_row=2, max_row=len(algorithms) + 1, min_col=1, max_col=1)
    uap_chart.add_data(uap_data_ref, titles_from_data=True)
    uap_chart.set_categories(uap_cats)
    
    # 设置图表大小
    uap_chart.width = 30
    uap_chart.height = 15
    
    # 添加图表到工作表
    chart_sheet.add_chart(uap_chart, "A20")

print("Excel文件已生成：g:/trae/myhash/scid_results.xlsx")

# 创建可视化图表
plt.figure(figsize=(15, 10))

# 绘制mAP图表
plt.subplot(2, 1, 1)
for i, algo in enumerate(algorithms):
    plt.plot(hash_sizes, map_data[i], marker='o', label=algo)
plt.xscale('log', base=2)
plt.xlabel('Hash Size')
plt.ylabel('mAP')
plt.title('mAP vs Hash Size')
plt.grid(True, alpha=0.3)
plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')

# 绘制μAP图表
plt.subplot(2, 1, 2)
for i, algo in enumerate(algorithms):
    plt.plot(hash_sizes, uap_data[i], marker='o', label=algo)
plt.xscale('log', base=2)
plt.xlabel('Hash Size')
plt.ylabel('μAP')
plt.title('μAP vs Hash Size')
plt.grid(True, alpha=0.3)
plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')

plt.tight_layout()
plt.savefig('g:/trae/myhash/scid_results_chart.png', dpi=300, bbox_inches='tight')
print("图表已保存：g:/trae/myhash/scid_results_chart.png")

# 创建三维柱状图功能

# 我将为您的代码添加三维柱状图功能，分别展示mAP和μAP的数据。以下是对代码的修改：

# 准备数据网格
X, Y = np.meshgrid(np.arange(len(hash_sizes)), np.arange(len(algorithms)))
hash_size_labels = [str(size) for size in hash_sizes]
algorithm_labels = [algo[:15] + '...' if len(algo) > 15 else algo for algo in algorithms]

# 创建mAP的三维柱状图
fig_map = plt.figure(figsize=(16, 10))
ax_map = fig_map.add_subplot(111, projection='3d')
dz = np.array(map_data).flatten()
z_zeros = np.zeros_like(dz)
dx = np.ones_like(dz) * 0.5
dy = np.ones_like(dz) * 0.5

# 设置颜色映射
colors = plt.cm.viridis(dz / np.max(dz))

# 绘制柱状图
ax_map.bar3d(X.flatten(), Y.flatten(), z_zeros, dx, dy, dz, color=colors, shade=True, alpha=0.8)

# 设置坐标轴标签和标题
ax_map.set_xlabel('Hash Size')
ax_map.set_ylabel('Algorithm')
ax_map.set_zlabel('mAP')
ax_map.set_title('3D Bar Chart of mAP Values')

# 设置坐标轴刻度标签
ax_map.set_xticks(np.arange(len(hash_sizes)))
ax_map.set_yticks(np.arange(len(algorithms)))
ax_map.set_xticklabels(hash_size_labels)
ax_map.set_yticklabels(algorithm_labels)

# 调整视角
ax_map.view_init(elev=30, azim=45)

# 保存图表
plt.tight_layout()
plt.savefig('g:/trae/myhash/scid_results_3d_map.png', dpi=300, bbox_inches='tight')
print("mAP三维柱状图已保存：g:/trae/myhash/scid_results_3d_map.png")

# 创建μAP的三维柱状图
fig_uap = plt.figure(figsize=(16, 10))
ax_uap = fig_uap.add_subplot(111, projection='3d')
dz = np.array(uap_data).flatten()
z_zeros = np.zeros_like(dz)
dx = np.ones_like(dz) * 0.5
dy = np.ones_like(dz) * 0.5

# 设置颜色映射
colors = plt.cm.viridis(dz / np.max(dz))

# 绘制柱状图
ax_uap.bar3d(X.flatten(), Y.flatten(), z_zeros, dx, dy, dz, color=colors, shade=True, alpha=0.8)

# 设置坐标轴标签和标题
ax_uap.set_xlabel('Hash Size')
ax_uap.set_ylabel('Algorithm')
ax_uap.set_zlabel('μAP')
ax_uap.set_title('3D Bar Chart of μAP Values')

# 设置坐标轴刻度标签
ax_uap.set_xticks(np.arange(len(hash_sizes)))
ax_uap.set_yticks(np.arange(len(algorithms)))
ax_uap.set_xticklabels(hash_size_labels)
ax_uap.set_yticklabels(algorithm_labels)

# 调整视角
ax_uap.view_init(elev=30, azim=45)

# 保存图表
plt.tight_layout()
plt.savefig('g:/trae/myhash/scid_results_3d_uap.png', dpi=300, bbox_inches='tight')
print("μAP三维柱状图已保存：g:/trae/myhash/scid_results_3d_uap.png")